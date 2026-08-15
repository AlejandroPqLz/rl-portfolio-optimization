"""Raw series data ingestion.

Downloads the four raw inputs of the historical dataset, stores them verbatim under
``data/raw/`` and records URL, timestamp and SHA-256 digest for each file in
``data/raw/provenance.json``.

Three commands, one entry point:

    uv run python -m portfolio_manager.download --refresh    download and write manifest
    uv run python -m portfolio_manager.download --verify     verify local files, no network
    uv run python -m portfolio_manager.download --report     first and last observation
"""

# Imports
# =====================================================================
import argparse
import hashlib
import io
import json
import re
import shutil
import sys
import tempfile
import urllib.request
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal

import openpyxl

# Constants
# =====================================================================
MANIFEST_NAME = "provenance.json"
MANIFEST_SCHEMA_VERSION = 1
USER_AGENT = "portfolio-manager/0.1 (TFM research ingestion)"
HTTP_TIMEOUT_SECONDS = 120
CHUNK_BYTES = 64 * 1024
DEFAULT_RAW_DIR = Path("data/raw")

TimeConvention = Literal["end_of_month", "monthly_average", "daily"]
ParserKind = Literal["ff_factors_zip", "fred_csv", "worldbank_xlsx"]
RawKind = Literal["monthly_return_pct", "annual_yield_pct", "price_usd"]

ZIP_MAGIC = b"PK\x03\x04"
HTML_MARKERS = (b"<!doctype html", b"<html")


# Source declarations
# =====================================================================
@dataclass(frozen=True)
class SourceSpec:
    """Declarative description of one raw input.

    Attributes:
        key: Stable logical name used in the manifest and downstream configuration.
        filename: Destination file name under the raw directory.
        url: Direct download URL.
        landing_url: Human-readable page documenting the series. Recorded because some
            direct URLs embed a rotating document identifier and stop resolving.
        description: What the file contains, for the provenance appendix.
        time_convention: How the monthly figure is built by the provider. Drives the
            limitation declared in the methodology chapter.
        parser: Format family. Selects the payload check and the coverage reader.
        raw_kind: Nature of the raw datum, to make the later conversion explicit.
    """

    key: str
    filename: str
    url: str
    landing_url: str
    description: str
    time_convention: TimeConvention
    parser: ParserKind
    raw_kind: RawKind


US_EQUITY = SourceSpec(
    key="us_equity_total_return",
    filename="us_equity_ff_factors.zip",
    url="https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_Factors_CSV.zip",
    landing_url="https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/data_library.html",
    description="Fama-French research factors, monthly. Total return of the US market portfolio "
    "is recovered downstream as Mkt-RF + RF.",
    time_convention="end_of_month",
    parser="ff_factors_zip",
    raw_kind="monthly_return_pct",
)

UST_10Y = SourceSpec(
    key="ust_10y_yield",
    filename="ust10y_dgs10.csv",
    url="https://fred.stlouisfed.org/graph/fredgraph.csv?id=DGS10",
    landing_url="https://fred.stlouisfed.org/series/DGS10",
    description="10-year Treasury constant maturity yield, daily, percent per annum.",
    time_convention="daily",
    parser="fred_csv",
    raw_kind="annual_yield_pct",
)

UST_2Y = SourceSpec(
    key="ust_2y_yield",
    filename="ust2y_dgs2.csv",
    url="https://fred.stlouisfed.org/graph/fredgraph.csv?id=DGS2",
    landing_url="https://fred.stlouisfed.org/series/DGS2",
    description="2-year Treasury constant maturity yield, daily, percent per annum.",
    time_convention="daily",
    parser="fred_csv",
    raw_kind="annual_yield_pct",
)

GOLD = SourceSpec(
    key="gold_usd",
    filename="gold_worldbank_pinksheet.xlsx",
    url="https://thedocs.worldbank.org/en/doc/74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/"
    "related/CMO-Historical-Data-Monthly.xlsx",
    landing_url="https://www.worldbank.org/en/research/commodity-markets",
    description="World Bank Commodity Markets monthly prices, sheet 'Monthly Prices', column 'Gold' "
    "in USD per troy ounce. The published monthly figure is the average of the month, not "
    "the closing value of the last trading day.",
    time_convention="monthly_average",
    parser="worldbank_xlsx",
    raw_kind="price_usd",
)

SOURCES: tuple[SourceSpec, ...] = (US_EQUITY, UST_10Y, UST_2Y, GOLD)


# Digest
# =====================================================================
def sha256_of(path: Path) -> str:
    """Return the hex SHA-256 digest of a file, read in chunks.

    Args:
        path: File to digest.
    """
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(CHUNK_BYTES), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _utc_now_iso() -> str:
    """Return the current UTC instant as an ISO 8601 string with second resolution."""
    return datetime.now(UTC).isoformat(timespec="seconds")


# Download
# =====================================================================
def assert_payload_shape(spec: SourceSpec, path: Path) -> None:
    """Reject a payload whose shape contradicts the declared format.

    Bot-protection challenges arrive with HTTP status 200 and an HTML body. Without this
    check such a page would be hashed and recorded as a valid series, and every later
    verification would pass because the HTML digest is stable.

    Args:
        spec: Declarative description of the source.
        path: Downloaded file to inspect.

    Raises:
        ValueError: If the payload is HTML or does not match the expected format.
    """
    head = path.read_bytes()[:512]
    lowered = head.lower()
    if any(marker in lowered for marker in HTML_MARKERS):
        raise ValueError(f"{spec.filename}: HTML received instead of data. The provider likely blocked the request.")
    if spec.parser in ("ff_factors_zip", "worldbank_xlsx") and not head.startswith(ZIP_MAGIC):
        raise ValueError(f"{spec.filename}: expected a zip container, found {head[:4]!r}")
    if spec.parser == "fred_csv" and b"," not in head.split(b"\n", 1)[0]:
        raise ValueError(f"{spec.filename}: first line has no comma, not a CSV")


def download_one(spec: SourceSpec, raw_dir: Path) -> dict[str, object]:
    """Download one source atomically and return its manifest record.

    The payload is written to a temporary file and moved into place only after the
    transfer completes, so an interrupted download never leaves a truncated file that
    would be hashed as if it were valid. The temporary file lives in the destination
    directory because ``Path.replace`` is atomic only within one filesystem.

    Args:
        spec: Declarative description of the source to download.
        raw_dir: Directory where the raw files are stored. Created if absent.

    Returns:
        A record with URL, HTTP status, timestamp, size and SHA-256 digest.

    Raises:
        urllib.error.URLError: On network or HTTP failure.
        ValueError: If the payload does not match the declared format.
    """
    raw_dir.mkdir(parents=True, exist_ok=True)
    destination = raw_dir / spec.filename
    request = urllib.request.Request(spec.url, headers={"User-Agent": USER_AGENT})

    with urllib.request.urlopen(request, timeout=HTTP_TIMEOUT_SECONDS) as response:
        raw_status = getattr(response, "status", None)
        status = int(raw_status) if raw_status is not None else None
        effective_url = response.geturl()
        last_modified = response.headers.get("Last-Modified")
        content_type = response.headers.get("Content-Type")

        with tempfile.NamedTemporaryFile(dir=raw_dir, delete=False) as staging:
            shutil.copyfileobj(response, staging)
            staging_path = Path(staging.name)

    try:
        assert_payload_shape(spec, staging_path)
    except ValueError:
        staging_path.unlink(missing_ok=True)
        raise

    staging_path.replace(destination)

    return {
        "key": spec.key,
        "filename": spec.filename,
        "description": spec.description,
        "raw_kind": spec.raw_kind,
        "time_convention": spec.time_convention,
        "url": spec.url,
        "effective_url": effective_url,
        "landing_url": spec.landing_url,
        "http_status": status,
        "content_type": content_type,
        "server_last_modified": last_modified,
        "downloaded_at_utc": _utc_now_iso(),
        "size_bytes": destination.stat().st_size,
        "sha256": sha256_of(destination),
    }


# Read and write manifest (provenance.json)
# =====================================================================
def write_manifest(records: list[dict[str, object]], raw_dir: Path) -> Path:
    """Serialise the manifest next to the raw files and return its path.

    Args:
        records: One record per source, in declaration order.
        raw_dir: Directory where the raw files are stored.

    Returns:
        Path to the written manifest.
    """
    manifest = {
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "generated_at_utc": _utc_now_iso(),
        "files": records,
    }
    path = raw_dir / MANIFEST_NAME
    path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return path


def read_manifest(raw_dir: Path) -> dict[str, object]:
    """Load the manifest.

    Args:
        raw_dir: Directory where the raw files are stored.

    Returns:
        The manifest as a dictionary.
    """
    path = raw_dir / MANIFEST_NAME
    if not path.is_file():
        raise FileNotFoundError(f"{path} not found. Run with --refresh first.")
    return json.loads(path.read_text(encoding="utf-8"))


def verify(raw_dir: Path) -> list[str]:
    """Check local files against the manifest without touching the network.

    Every problem is collected instead of raised, so a single run reports all the files
    that need attention rather than only the first one.

    Args:
        raw_dir: Directory where the raw files are stored.

    Returns:
        Human-readable problems. An empty list means every declared file is present and
        byte-identical to the one recorded at download time.
    """
    manifest = read_manifest(raw_dir)
    files = manifest.get("files", [])
    if not isinstance(files, list):
        return [f"malformed manifest: 'files' is {type(files).__name__}, expected list"]
    problems: list[str] = []
    for record in files:
        filename = str(record["filename"])
        path = raw_dir / filename
        if not path.is_file():
            problems.append(f"{filename}: missing. Retrieve it from {record['url']}")
            continue
        actual = sha256_of(path)
        expected = str(record["sha256"])
        if actual != expected:
            problems.append(f"{filename}: digest mismatch. expected {expected[:16]}…, found {actual[:16]}…")
    return problems


# Coverage report, read-only
# =====================================================================
def _observation_range_ff(path: Path) -> tuple[str, str, int]:
    """First and last monthly label of the Fama-French factor file.

    The monthly block ends where the annual block begins. Reading past that boundary
    mixes annual with monthly observations, which produces no error and invalid figures.
    """
    with zipfile.ZipFile(path) as archive:
        name = next(n for n in archive.namelist() if n.upper().endswith(".CSV"))
        text = archive.read(name).decode("latin-1")
    labels: list[str] = []
    for line in io.StringIO(text):
        if "Annual Factors" in line:
            break
        token = line.split(",", 1)[0].strip()
        if re.fullmatch(r"\d{6}", token):
            labels.append(token)
    return labels[0], labels[-1], len(labels)


def _observation_range_fred(path: Path) -> tuple[str, str, int]:
    """First and last dated row of a FRED CSV, ignoring the holiday marker '.'."""
    dates: list[str] = []
    with path.open(encoding="utf-8-sig") as handle:
        next(handle, None)
        for line in handle:
            parts = line.strip().split(",")
            if len(parts) < 2 or parts[1].strip() in {"", "."}:
                continue
            dates.append(parts[0].strip())
    return dates[0], dates[-1], len(dates)


def _observation_range_worldbank(path: Path) -> tuple[str, str, int]:
    """First and last monthly label of the World Bank monthly prices sheet.

    The gold column is located by name because its position shifts when new commodities
    are appended to the publication.
    """
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book["Monthly Prices"]
        header = next(sheet.iter_rows(min_row=5, max_row=5, values_only=True))
        gold_index = next((i for i, cell in enumerate(header) if cell and str(cell).strip().lower() == "gold"), None)
        if gold_index is None:
            raise RuntimeError("column 'Gold' not found in sheet 'Monthly Prices'")
        labels: list[str] = []
        for row in sheet.iter_rows(min_row=7, values_only=True):
            label, value = row[0], row[gold_index]
            if label is None or value is None or isinstance(value, str):
                continue
            labels.append(str(label).strip())
    finally:
        book.close()
    return labels[0], labels[-1], len(labels)


_RANGE_READERS = {
    "ff_factors_zip": _observation_range_ff,
    "fred_csv": _observation_range_fred,
    "worldbank_xlsx": _observation_range_worldbank,
}


def coverage_report(raw_dir: Path) -> list[tuple[str, str]]:
    """Report first and last observation per file, read-only.

    A diagnostic, not a transformation: nothing is written and no value is converted.
    It exists so the window decision is taken against an observed date rather than an
    assumption. An unreadable file is reported instead of aborting the whole report.

    Args:
        raw_dir: Directory where the raw files are stored.

    Returns:
        One pair of source key and detail line per declared source.
    """
    lines: list[tuple[str, str]] = []
    for spec in SOURCES:
        path = raw_dir / spec.filename
        if not path.is_file():
            lines.append((spec.key, "missing"))
            continue
        try:
            first, last, count = _RANGE_READERS[spec.parser](path)
        except Exception as error:  # pylint: disable=broad-except
            lines.append((spec.key, f"unreadable: {error}"))
            continue
        lines.append((spec.key, f"{first} … {last}   n={count}   convention={spec.time_convention}"))
    return lines


# Entry point
# =====================================================================
def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python -m portfolio_manager.download",
        description="Download the raw historical series and record their provenance.",
    )
    parser.add_argument("--refresh", action="store_true", help="download from the network and rewrite the manifest")
    parser.add_argument("--verify", action="store_true", help="check local files against the manifest, no network")
    parser.add_argument("--report", action="store_true", help="print first and last observation of each raw file")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Command line entry point.

    Verification runs after ``--refresh`` and whenever no action is requested, so the
    bare command is a safe integrity check.

    Args:
        argv: Argument list, defaulting to ``sys.argv[1:]``.

    Returns:
        Process exit code: 0 when every declared file matches the manifest.
    """
    args = _build_parser().parse_args(argv)
    raw_dir: Path = DEFAULT_RAW_DIR

    if args.refresh:
        records: list[dict[str, object]] = []
        for spec in SOURCES:
            try:
                record = download_one(spec, raw_dir)
            except (OSError, ValueError) as error:
                print(f"FAIL: {error}", file=sys.stderr)
                return 1
            records.append(record)
            print(f"downloaded  {spec.filename}  {record['size_bytes']} B  {str(record['sha256'])[:16]}…")
        print(f"manifest {write_manifest(records, raw_dir)}")

    if args.refresh or args.verify or not args.report:
        try:
            problems = verify(raw_dir)
        except FileNotFoundError as error:
            print(f"FAIL: {error}", file=sys.stderr)
            return 1
        for problem in problems:
            print(f"FAIL: {problem}", file=sys.stderr)
        if problems:
            return 1
        print(f"verified {len(SOURCES)} files match the manifest")

    if args.report:
        print("\ncoverage report (read-only)")
        for key, detail in coverage_report(raw_dir):
            print(f"  {key:<24} {detail}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
